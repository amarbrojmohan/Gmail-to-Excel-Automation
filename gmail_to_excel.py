from __future__ import print_function
import os.path
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from openpyxl import Workbook, load_workbook

# ------------------------------------------------------------
# Configuration
# ------------------------------------------------------------
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']
EXCEL_FILE = "gmail_emails.xlsx"

def get_gmail_service():
    """Authenticate with Gmail and return a service object."""
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('gmail', 'v1', credentials=creds)

def fetch_emails(service, max_results=50):
    """Fetch basic metadata (subject, from, date) for recent messages."""
    results = service.users().messages().list(userId='me', maxResults=max_results).execute()
    messages = results.get('messages', [])
    data = []
    for msg in messages:
        msg_id = msg['id']
        m = service.users().messages().get(
            userId='me',
            id=msg_id,
            format='metadata',
            metadataHeaders=['Subject', 'From', 'Date']
        ).execute()

        headers = m['payload']['headers']
        subject = sender = date = ''
        for h in headers:
            name, value = h['name'], h['value']
            if name == 'Subject':
                subject = value
            elif name == 'From':
                sender = value
            elif name == 'Date':
                date = value
        data.append((msg_id, sender, subject, date))
    return data

def load_existing_ids():
    """Read message IDs from Excel to prevent duplicates."""
    if not os.path.exists(EXCEL_FILE):
        return set()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ids = {row[0].value for row in ws.iter_rows(min_row=2, max_col=1) if row[0].value}
    return ids

def write_to_excel(data):
    """Write new email records into Excel, skipping duplicates."""
    existing_ids = load_existing_ids()

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Message ID", "From", "Subject", "Date"])
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    new_count = 0
    for row in data:
        msg_id, sender, subject, date = row
        if msg_id not in existing_ids:
            ws.append(row)
            existing_ids.add(msg_id)
            new_count += 1

    wb.save(EXCEL_FILE)
    print(f"✅ Added {new_count} new email(s) to {EXCEL_FILE}. (Total tracked: {len(existing_ids)})")

def main():
    service = get_gmail_service()
    emails = fetch_emails(service, max_results=50)
    write_to_excel(emails)

if __name__ == '__main__':
    main()
